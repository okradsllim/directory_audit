import pandas as pd
from pathlib import Path
from shutil import move
import logging
import os
from openpyxl import load_workbook
import re

def extract_hyperlinks(excel_file_path, sheet_name='AuditSheet', target_column_name='Path'):
    wb = load_workbook(excel_file_path, data_only=False)
    sheet = wb[sheet_name]

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        target_column_index = header_row.index(target_column_name)
    except ValueError:
        logging.error(f"Column '{target_column_name}' not found in the sheet '{sheet_name}'.")
        return {}

    hyperlink_regex = re.compile(r'HYPERLINK\("([^"]+)"')
    hyperlinks = {}
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        cell = row[target_column_index]
        if cell.value and "HYPERLINK" in str(cell.value):
            match = hyperlink_regex.search(cell.value)
            if match:
                link_location = match.group(1)
                hyperlinks[row_index] = link_location
                logging.debug(f"Hyperlink extracted at row {row_index}: {link_location}")
            else:
                logging.debug(f"Hyperlink formula not found at row {row_index}")
                hyperlinks[row_index] = None
        else:
            logging.debug(f"No hyperlink at row {row_index}")
            hyperlinks[row_index] = None

    return hyperlinks

def extract_base_directory(paths):
    absolute_paths = [Path(p).resolve() for p in paths if p is not None]
    if not absolute_paths:
        return None
    common_path = os.path.commonpath(absolute_paths)
    return Path(common_path)

def action_rename(original_path, new_name):
    logging.debug(f"Attempting to rename: {original_path} to {new_name}")

    if not original_path.exists():
        log_message = f"Rename Failed - Original file/folder does not exist: {original_path}"
        logging.warning(log_message)
        return log_message

    new_path = original_path.parent / new_name
    
    if original_path.is_file(): # If the original path is a file, preserve the extension in the new name
        original_extension = original_path.suffix
        if not new_name.endswith(original_extension):
            new_name += original_extension

    if new_path.exists():
        log_message = f"Rename Failed - New name already exists: {new_path}"
        logging.warning(log_message)
        return log_message

    try:
        original_path.rename(new_path)
        log_message = f"Renamed {original_path} to {new_path}"
        logging.info(log_message)
        print(log_message)
        return new_name  # Return the full new_name including the extension
    except Exception as e:
        log_message = f"Rename Error: {e}"
        logging.error(log_message)
        return log_message

def action_move(original_path, move_to_folder_name, folders_to_create, BASE_DIRECTORY):
    logging.debug(f"Attempting to move: {original_path} to {move_to_folder_name}")

    # Retrieve the target directory from the folders_to_create dictionary
    target_dir = folders_to_create.get(move_to_folder_name)

    if not original_path.exists():
        log_message = f"Move Failed - Original file/folder does not exist: {original_path}"
        logging.warning(log_message)
        return log_message

    if target_dir is None:
        log_message = f"Move Failed - Target directory not found for: {move_to_folder_name}"
        logging.warning(log_message)
        return log_message

    if not target_dir.is_dir():
        log_message = f"Move Failed - Target directory is not a directory: {target_dir}"
        logging.warning(log_message)
        return log_message

    if original_path.resolve() == target_dir.resolve():
        log_message = "Move Skipped - Source and target are identical"
        logging.info(log_message)
        return log_message

    try:
        print(f"Attempting to move {original_path} to {target_dir}")
        move(str(original_path), str(target_dir))
        log_message = f"Moved {original_path} to {target_dir}"
        logging.info(log_message)
        print(log_message)
        return log_message
    
    except Exception as e:
        log_message = f"Move Error: {e}"
        logging.error(log_message)
        print(log_message)
        return log_message

def action_delete(original_path, recycle_dir_path):
    
    logging.debug(f"Attempting to move: {original_path} to {recycle_dir_path}")
    
    if original_path.exists():
        try:
            move(str(original_path), str(recycle_dir_path))
            log_message = f'Moved {original_path} to recycle directory: {recycle_dir_path}'
            logging.info(log_message)
            print(log_message)
            return log_message
        except Exception as e:
            log_message = f"Delete Error: {e}"
            logging.error(log_message)
            return log_message
    else:
        log_message = 'Delete Failed - File not found'
        logging.warning(log_message)
        return log_message

def perform_actions(audit_sheet_df, recycle_dir_path, folders_to_create, action_logs, BASE_DIRECTORY):
    
    logging.debug("Starting to perform actions.")

    for index, row in audit_sheet_df.iterrows():
        logging.debug(f"Processing row {index + 1}.")
        name = row['Name']
        actions = row['Action'].lower().split(',') if isinstance(row['Action'], str) else []
        new_name = row['Rename as…'].strip() if isinstance(row['Rename as…'], str) else ''
        move_to = row['Move to…'].strip() if isinstance(row['Move to…'], str) else ''

        original_path = Path(row['Extracted Path'])
        
        # Check if the path points to an existing file or directory
        if not original_path.exists():
            action_logs.append({'Action': ', '.join(actions), 'Path': str(original_path), 'Status': 'Original path not found or does not exist'})
            continue

        status = []
        for action in actions:
            logging.debug(f"Performing action '{action}' for {name}.")
            if action == 'rename' and new_name:
                result = action_rename(original_path, new_name)
                status.append(result)
                if not result.startswith("Failed"):
                    original_path = Path(result)
                logging.info(status[-1])

            elif action == 'move' and move_to:
                result = action_move(original_path, move_to, folders_to_create, BASE_DIRECTORY)
                status.append(result) 
                logging.info(status[-1])
                
            elif action == 'delete':
                result = action_delete(original_path, recycle_dir_path)
                status.append(result)
                logging.info(status[-1])
                
            else:
                status.append(f"Failed - Target directory not found for moving: {move_to}")
                logging.info(status[-1])

        action_logs.append({'Action': ', '.join(actions), 'Path': str(original_path), 'Status': '; '.join(status)})
        
        logging.debug(f"Row {index + 1} actions completed with status: {'; '.join(status)}")
    
def validate_path(path, should_exist=True, is_directory=False):
    p = Path(path).resolve()
    if should_exist and not p.exists():
        print(f"Path does not exist: {p}")
        return False
    if is_directory and not p.is_dir():
        print(f"Path is not a directory: {p}")
        return False
    return True

def get_validated_path(prompt, should_exist=True, is_directory=False, max_attempts=5):
    attempts = 0
    while attempts < max_attempts:
        user_input = input(f"{prompt} (Type 'exit' to quit, {max_attempts - attempts} attempts left): ").strip()
        
        if user_input.lower() == 'exit':
            print("Exiting program.")
            return None

        path = Path(user_input).resolve()
        if validate_path(path, should_exist, is_directory):
            return path

        attempts += 1
        print(f"Invalid input. Please try again.")

    print("Maximum number of attempts reached. Exiting program.")
    return None

def main():

    desktop_dir = os.path.join(os.path.expanduser('~'), 'Desktop')
    
    log_file_path = os.path.join(desktop_dir, 'file_manager.log')
    logging.basicConfig(filename=log_file_path, level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
    
    action_logs = []

    input_excel_file = get_validated_path("Enter path to the Excel file: ", should_exist=True, is_directory=False)
    if input_excel_file is None:
        return 
    
    audit_sheet_df = pd.read_excel(input_excel_file, sheet_name='AuditSheet')
    
    ## Initial Check for Actionable Data
    if audit_sheet_df['Action'].isna().all():
        print("No actions specified in the 'Action' column. Exiting program.")
        return
    
    ## Normalize and Filter Actions
    valid_actions = ['delete', 'rename', 'move']
    audit_sheet_df['Action'] = audit_sheet_df['Action'].fillna('').astype(str).str.lower().str.strip()
    valid_action_rows = audit_sheet_df[audit_sheet_df['Action'].str.contains('|'.join(valid_actions))]
    
    ## Further Validation of Actions
    if valid_action_rows.empty:
        print("No valid actions (delete, rename, move) found in the 'Action' column. Exiting program.")
        return

    # Extract Hyperlinks
    hyperlinks = extract_hyperlinks(input_excel_file)
    valid_action_rows.loc[:, 'Extracted Path'] = valid_action_rows.index + 2
    valid_action_rows.loc[:, 'Extracted Path'] = valid_action_rows['Extracted Path'].apply(lambda x: hyperlinks.get(x))
    
    # Determine the Base Directory
    BASE_DIRECTORY = extract_base_directory(valid_action_rows['Extracted Path'])
    print(f"Base directory: {BASE_DIRECTORY}")
                
    move_to_folders = audit_sheet_df[audit_sheet_df['Action'].str.contains('move', na=False)]['Move to…'].unique()
    folders_to_create = {}

    for folder_name in move_to_folders:
        found_folder_path = None
        
        # Check in subdirectories of BASE_DIRECTORY and its siblings
        potential_directories = [BASE_DIRECTORY] + [x for x in BASE_DIRECTORY.parent.iterdir() if x.is_dir()]
        for dir in potential_directories:
            potential_folder_path = dir / folder_name
            if potential_folder_path.exists():
                found_folder_path = potential_folder_path
                break

        # Iterate through each parent of the BASE_DIRECTORY in reverse order
        if found_folder_path is None:
            for parent in reversed(BASE_DIRECTORY.parents):
                potential_folder_path = parent / folder_name
                if potential_folder_path.exists():
                    found_folder_path = potential_folder_path
                    break
            
        # Create the folder as a sibling of BASE_DIRECTORY if not found
        if found_folder_path is None:
            folder_path_to_create = BASE_DIRECTORY.parent / folder_name
            folder_path_to_create.mkdir(parents=True, exist_ok=True)
            logging.info(f"Directory created: {folder_path_to_create}")
            print(f"Directory created: {folder_path_to_create}")
            folders_to_create[folder_name] = folder_path_to_create
        else:
            logging.info(f"Folder found: {found_folder_path}")
            folders_to_create[folder_name] = found_folder_path


    if audit_sheet_df['Action'].str.contains(r'delete', case=False, na=False).any():
        while True:
            recycle_dir_input = input("Enter the path to the recycle directory: ").strip()
            recycle_dir_path = Path(recycle_dir_input).resolve()

            # Check if the recycle directory is not inside BASE_DIRECTORY
            if BASE_DIRECTORY in recycle_dir_path.parents:
                print(f"The recycle directory should not be inside the base directory '{BASE_DIRECTORY}'. Please choose a different location.")
                continue

            
            if not recycle_dir_path.exists():
                create_dir = input(f"The path {recycle_dir_path} does not exist. Do you want to create it? (yes/no): ").strip().lower()
                if create_dir == 'yes':
                    recycle_dir_path.mkdir(parents=True, exist_ok=True)
                    print(f"Directory created: {recycle_dir_path}")
                    break
                elif create_dir == 'no':
                    continue
                else:
                    print("Invalid input. Please type 'yes' or 'no'.")
            elif not recycle_dir_path.is_dir():
                print(f"The path is not a directory: {recycle_dir_path}. Please enter a valid directory path.")
                continue
            else:
                break
    else:
        recycle_dir_path = None 

    # Performing actions
    perform_actions(valid_action_rows, recycle_dir_path, folders_to_create, action_logs, BASE_DIRECTORY)
    
    for log in action_logs:
        print(f"Action: {log['Action']}, Path: {log['Path']}, Status: {log['Status']}")

if __name__ == "__main__":
    main()